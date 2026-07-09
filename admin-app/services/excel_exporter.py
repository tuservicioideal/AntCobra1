"""
Excel Exporter Service
=====================

Exports client management data (gestión results) to an Excel file
matching the bank's required output format.

Output columns:
  A: Etapa               — etapa_deuda
  B: Nombre proveedor    — configurable (default "PERECAUDOL")
  C: Numero campaña      — campana_banco
  D: Codigo cliente      — codigo_cliente
  E: Fecha gestion       — date part of fecha_gestion
  F: Hora gestion        — time part of fecha_gestion
  G: Nivel 1             — nivel_1
  H: Nivel 2             — nivel_2
  I: Nivel 3             — nivel_3
  J: Nivel 4             — nivel_4
  K: Fecha promesa pago  — fecha_promesa_pago
  L: Monto promesa pago  — monto_promesa_pago
  M: Observación         — nota_gestor
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Column definitions ──────────────────────────────────────────

EXPORT_COLUMNS = [
    {"key": "etapa_deuda",        "header": "Etapa",              "width": 10},
    {"key": "nombre_proveedor",   "header": "Nombre proveedor",   "width": 22},
    {"key": "campana_banco",      "header": "Numero campaña",     "width": 16},
    {"key": "codigo_cliente",     "header": "Codigo cliente",     "width": 16},
    {"key": "_fecha_gestion",     "header": "Fecha gestion",      "width": 14},
    {"key": "_hora_gestion",      "header": "Hora gestion",       "width": 12},
    {"key": "nivel_1",            "header": "Nivel 1",            "width": 26},
    {"key": "nivel_2",            "header": "Nivel 2",            "width": 32},
    {"key": "nivel_3",            "header": "Nivel 3",            "width": 38},
    {"key": "nivel_4",            "header": "Nivel 4",            "width": 42},
    {"key": "fecha_promesa_pago", "header": "Fecha promesa pago", "width": 18},
    {"key": "monto_promesa_pago", "header": "Monto promesa pago", "width": 20},
    {"key": "nota_gestor",        "header": "Observación",        "width": 30},
]


def _parse_fecha_hora(raw_fecha_gestion: Any) -> tuple[str, str]:
    """Extract date and time strings from a fecha_gestion value.

    Handles:
      - datetime objects
      - ISO-format strings ("2023-03-07 12:03:49", "2023-03-07T12:03:49")
      - Firestore Timestamp objects (with isoformat())
      - Empty / None

    Returns:
        (date_str "dd/mm/yyyy", time_str "HH:MM:SS") or ("", "")
    """
    if not raw_fecha_gestion:
        return ("", "")

    dt_obj = None

    if isinstance(raw_fecha_gestion, datetime):
        dt_obj = raw_fecha_gestion
    elif hasattr(raw_fecha_gestion, "isoformat"):
        # Firestore Timestamp or date objects
        try:
            dt_obj = raw_fecha_gestion
        except Exception:
            pass
    elif isinstance(raw_fecha_gestion, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S.%f",
                    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                dt_obj = datetime.strptime(raw_fecha_gestion.strip(), fmt)
                break
            except ValueError:
                continue

    if dt_obj is not None:
        try:
            date_str = dt_obj.strftime("%d/%m/%Y")
            time_str = dt_obj.strftime("%H:%M:%S")
            return (date_str, time_str)
        except Exception:
            pass

    # Fall back: return raw as date, empty time
    return (str(raw_fecha_gestion), "")


def export_gestion_excel(
    clientes: List[Dict[str, Any]],
    output_path: str,
    nombre_proveedor: str = "PERECAUDOL",
    solo_gestionados: bool = True,
) -> str:
    """
    Export management results to an Excel file.

    Args:
        clientes: List of client dicts (from campaign_manager.get_all_clients).
        output_path: Full path for the output .xlsx file.
        nombre_proveedor: Provider name to fill in column B.
        solo_gestionados: If True, skip clients still in "pendiente" status.

    Returns:
        The path to the generated Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestión"

    # ── Styles ───────────────────────────────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=10)
    money_fmt = '#,##0'

    # ── Headers (row 1) ─────────────────────────────────────
    for col_idx, col_def in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    # ── Data rows ────────────────────────────────────────────
    row_num = 2
    for client in clientes:
        estado = client.get("estado_gestion", "pendiente")
        if solo_gestionados and estado == "pendiente":
            continue

        fecha_str, hora_str = _parse_fecha_hora(client.get("fecha_gestion"))

        row_data = []
        for col_def in EXPORT_COLUMNS:
            key = col_def["key"]
            if key == "nombre_proveedor":
                row_data.append(nombre_proveedor)
            elif key == "_fecha_gestion":
                row_data.append(fecha_str)
            elif key == "_hora_gestion":
                row_data.append(hora_str)
            elif key == "monto_promesa_pago":
                val = client.get(key, 0)
                row_data.append(float(val) if val else None)
            else:
                row_data.append(client.get(key, ""))

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if EXPORT_COLUMNS[col_idx - 1]["key"] == "monto_promesa_pago" and value:
                cell.number_format = money_fmt

        row_num += 1

    # ── Freeze header row ────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ──────────────────────────────────────────
    last_col = get_column_letter(len(EXPORT_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{row_num - 1}"

    # ── Save ─────────────────────────────────────────────────
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
