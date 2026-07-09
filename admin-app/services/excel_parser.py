"""
Excel Parser Service
Reads and processes Excel files from the bank, extracting key client data
and grouping them by Seccion (gestor assignment).
"""

import openpyxl
from datetime import datetime
from collections import defaultdict
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import EXCEL_COLUMNS


def safe_str(value):
    """Convert a value to string safely, handling None and dates."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def safe_float(value):
    """Convert a value to float safely."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def safe_int(value):
    """Convert a value to int safely."""
    if value is None:
        return 0
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def make_seccion_key(region: str, zona: str, seccion: str) -> str:
    """Build composite key: region_zona_seccion (e.g. '01_1211_H').

    This ensures uniqueness when the same section letter appears in
    different region/zona combinations.
    """
    r = str(region).strip() or "SR"
    z = str(zona).strip() or "SZ"
    s = str(seccion).strip().upper() or "SS"
    return f"{r}_{z}_{s}"


def parse_excel(file_path: str) -> dict:
    """
    Parse an Excel file and return structured data.
    
    Returns:
        dict with keys:
            - 'all_clients': list of all client dicts
            - 'by_seccion': dict grouping clients by their Seccion letter
            - 'summary': dict with summary statistics
            - 'headers': list of original column headers
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Use first sheet
    ws = wb[wb.sheetnames[0]]
    
    # Read headers from first row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(safe_str(cell.value))
    
    all_clients = []
    by_seccion = defaultdict(list)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[EXCEL_COLUMNS["codigo_cliente"]] is None:
            continue  # Skip empty rows
        
        client = {}
        for field_name, col_index in EXCEL_COLUMNS.items():
            if col_index < len(row):
                raw_value = row[col_index]
                
                # Apply type conversion based on field
                if field_name in ("dias_atraso", "edad"):
                    client[field_name] = safe_int(raw_value)
                elif field_name.startswith("importe_") or field_name.startswith("coordenada_"):
                    client[field_name] = safe_float(raw_value)
                elif field_name.startswith("fecha_"):
                    client[field_name] = safe_str(raw_value)
                else:
                    client[field_name] = safe_str(raw_value)
            else:
                client[field_name] = ""
        
        # Build full name for convenience
        client["nombre_completo"] = f"{client['nombres']} {client['apellido_paterno']} {client['apellido_materno']}".strip()
        
        all_clients.append(client)
        seccion = client.get("seccion", "SIN_SECCION")
        if not seccion:
            seccion = "SIN_SECCION"
        # Normalize section to uppercase for consistent matching with user profiles
        seccion = str(seccion).strip().upper()
        client["seccion"] = seccion

        # Composite key ensures uniqueness across region/zona combos
        region_val = client.get("region", "")
        zona_val = client.get("zona", "")
        sec_key = make_seccion_key(region_val, zona_val, seccion)
        client["seccion_key"] = sec_key
        by_seccion[sec_key].append(client)
    
    # Build summary
    total_deuda_asignada = sum(c.get("importe_deuda_asignada", 0) for c in all_clients)
    total_deuda_pendiente = sum(c.get("importe_deuda_pendiente", 0) for c in all_clients)
    
    summary = {
        "total_clientes": len(all_clients),
        "total_secciones": len(by_seccion),
        "secciones": {k: len(v) for k, v in sorted(by_seccion.items())},
        "total_deuda_asignada": round(total_deuda_asignada, 2),
        "total_deuda_pendiente": round(total_deuda_pendiente, 2),
        "departamentos": list(set(c.get("departamento", "") for c in all_clients if c.get("departamento"))),
    }
    
    wb.close()
    
    return {
        "all_clients": all_clients,
        "by_seccion": dict(by_seccion),
        "summary": summary,
        "headers": headers,
    }


def get_seccion_summary(by_seccion: dict) -> list:
    """
    Get a summary table for each seccion/gestor.
    
    Keys in *by_seccion* are composite keys (region_zona_seccion).
    Returns list of dicts with seccion stats.
    """
    result = []
    for seccion_key in sorted(by_seccion.keys()):
        clients = by_seccion[seccion_key]
        deuda_total = sum(safe_float(c.get("importe_deuda_asignada", 0)) for c in clients)
        deuda_pendiente = sum(safe_float(c.get("importe_deuda_pendiente", 0)) for c in clients)
        departamentos = list(set(c.get("departamento", "") for c in clients if c.get("departamento")))

        # Extract plain letter from composite key (last segment after underscore)
        parts = seccion_key.rsplit("_", 1)
        letra = parts[-1] if parts else seccion_key

        result.append({
            "seccion": seccion_key,  # composite key (used as map key)
            "seccion_letra": letra,  # plain letter for display badges
            "num_clientes": len(clients),
            "deuda_asignada": round(deuda_total, 2),
            "deuda_pendiente": round(deuda_pendiente, 2),
            "departamentos": ", ".join(departamentos),
        })
    
    return result


def get_hierarchy(all_clients: list) -> dict:
    """
    Build the Region → Zona → Sección hierarchy from parsed client data.

    Returns:
        {
            "regions": {
                "01": {
                    "zonas": {
                        "1211": {
                            "secciones": {
                                "H": {"num_clientes": 50, "deuda_asignada": ..., "deuda_pendiente": ...},
                            },
                            "num_clientes": ..., "deuda_asignada": ..., "deuda_pendiente": ...,
                        },
                    },
                    "num_clientes": ..., "deuda_asignada": ..., "deuda_pendiente": ...,
                },
            },
            "totals": {"num_clientes": ..., "deuda_asignada": ..., "deuda_pendiente": ...},
        }
    """
    regions: dict = {}

    for c in all_clients:
        region = safe_str(c.get("region")) or "SIN_REGION"
        zona = safe_str(c.get("zona")) or "SIN_ZONA"
        seccion = safe_str(c.get("seccion")) or "SIN_SECCION"
        deuda_a = safe_float(c.get("importe_deuda_asignada", 0))
        deuda_p = safe_float(c.get("importe_deuda_pendiente", 0))

        if region not in regions:
            regions[region] = {"zonas": {}, "num_clientes": 0,
                               "deuda_asignada": 0.0, "deuda_pendiente": 0.0}
        r = regions[region]
        r["num_clientes"] += 1
        r["deuda_asignada"] += deuda_a
        r["deuda_pendiente"] += deuda_p

        if zona not in r["zonas"]:
            r["zonas"][zona] = {"secciones": {}, "num_clientes": 0,
                                "deuda_asignada": 0.0, "deuda_pendiente": 0.0}
        z = r["zonas"][zona]
        z["num_clientes"] += 1
        z["deuda_asignada"] += deuda_a
        z["deuda_pendiente"] += deuda_p

        if seccion not in z["secciones"]:
            z["secciones"][seccion] = {"num_clientes": 0,
                                       "deuda_asignada": 0.0, "deuda_pendiente": 0.0}
        s = z["secciones"][seccion]
        s["num_clientes"] += 1
        s["deuda_asignada"] += deuda_a
        s["deuda_pendiente"] += deuda_p

    total_c = sum(r["num_clientes"] for r in regions.values())
    total_da = sum(r["deuda_asignada"] for r in regions.values())
    total_dp = sum(r["deuda_pendiente"] for r in regions.values())

    return {
        "regions": dict(sorted(regions.items())),
        "totals": {
            "num_clientes": total_c,
            "deuda_asignada": round(total_da, 2),
            "deuda_pendiente": round(total_dp, 2),
        },
    }


if __name__ == "__main__":
    # Quick test
    import json
    data = parse_excel(r"E:\FYM\PROY\AntonioCobranzas\AntCobra1\MUESTRA DE DATA.xlsx")
    print(f"Total clientes: {data['summary']['total_clientes']}")
    print(f"Secciones: {data['summary']['secciones']}")
    print(f"Deuda asignada total: S/ {data['summary']['total_deuda_asignada']}")
    print(f"Deuda pendiente total: S/ {data['summary']['total_deuda_pendiente']}")
    print(f"\nResumen por gestor:")
    for s in get_seccion_summary(data['by_seccion']):
        print(f"  Sección {s['seccion']}: {s['num_clientes']} clientes, Deuda: S/ {s['deuda_asignada']}, Pendiente: S/ {s['deuda_pendiente']}")
