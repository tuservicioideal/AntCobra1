"""
Functional validation tests for AntCobranzas admin-app.
Tests all new services, database migration, and UI integrations.
"""
import sys
import os
import json
import tempfile
import traceback
from datetime import date, datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

PASS = 0
FAIL = 0

def test(name, fn):
    global PASS, FAIL
    try:
        fn()
        PASS += 1
        print(f"  [PASS] {name}")
    except Exception as e:
        FAIL += 1
        print(f"  [FAIL] {name}: {e}")
        traceback.print_exc()

def fresh_db():
    """Create a fresh DatabaseService with a temp DB."""
    from services.database import DatabaseService
    tmp = tempfile.mktemp(suffix=".db")
    svc = DatabaseService(db_path=tmp)
    svc.initialize()
    return svc, tmp

def cleanup_db(svc, tmp):
    """Dispose engine and remove temp DB file."""
    try:
        if svc.engine:
            svc.engine.dispose()
    except Exception:
        pass
    try:
        if os.path.exists(tmp):
            os.unlink(tmp)
        # Also WAL/SHM files
        for ext in ("-wal", "-shm"):
            p = tmp + ext
            if os.path.exists(p):
                os.unlink(p)
    except Exception:
        pass

# ═══════════════════════════════════════════════════════════════
# 1. DATABASE SCHEMA V2
# ═══════════════════════════════════════════════════════════════
print("\n=== 1. DATABASE SCHEMA V2 ===")

def test_db_imports():
    from services.database import (
        db_service, Cliente, Campana, SyncLog,
        CURRENT_SCHEMA_VERSION, EstadoGestion, EstadoCampana,
    )
    assert CURRENT_SCHEMA_VERSION == 2

test("Database imports & schema version", test_db_imports)

def test_db_init_and_schema():
    svc, tmp = fresh_db()

    # Verify new columns exist on Cliente
    from sqlalchemy import inspect
    insp = inspect(svc.engine)
    cols = {c["name"] for c in insp.get_columns("clientes")}
    for col in ["nivel_1", "nivel_2", "nivel_3", "nivel_4",
                "canal_gestion", "fecha_promesa_pago", "monto_promesa_pago"]:
        assert col in cols, f"Missing column: {col}"

    # Verify sync_log table exists
    tables = insp.get_table_names()
    assert "sync_log" in tables, "Missing table: sync_log"

    # Verify SyncLog columns
    sync_cols = {c["name"] for c in insp.get_columns("sync_log")}
    for col in ["id", "tipo", "fecha", "registros_afectados", "resultado", "detalle"]:
        assert col in sync_cols, f"Missing sync_log column: {col}"

    cleanup_db(svc, tmp)

test("DB init with new schema columns", test_db_init_and_schema)

def test_schema_v16_historial_reparto_call():
    """Schema v16: historial_reparto_call table for call distribution audit."""
    from sqlalchemy import inspect
    from services.database import SchemaVersion, HistorialRepartoCall, CURRENT_SCHEMA_VERSION
    svc, tmp = fresh_db()
    insp = inspect(svc.engine)
    assert "historial_reparto_call" in insp.get_table_names()
    cols = {c["name"] for c in insp.get_columns("historial_reparto_call")}
    for col in ("tipo", "motivo", "detalle_json", "firebase_ok", "algoritmo"):
        assert col in cols, f"Missing historial_reparto_call column: {col}"
    with svc.session() as session:
        sv = session.query(SchemaVersion).first()
        assert sv is not None
        assert sv.version >= 16
        assert CURRENT_SCHEMA_VERSION >= 16
        session.add(HistorialRepartoCall(
            campana_id="test",
            tipo="reparto_inicial",
            motivo="test",
            algoritmo="LPT",
            cuentas_afectadas=1,
            monto_afectado=100.0,
            detalle_json="{}",
            firebase_ok=True,
        ))
        session.commit()
    cleanup_db(svc, tmp)

test("Schema v16 historial_reparto_call", test_schema_v16_historial_reparto_call)

def test_db_cliente_nivel_fields():
    """Test that Cliente ORM can handle nivel fields."""
    from services.database import Cliente, Campana, EstadoCampana
    svc, tmp = fresh_db()

    with svc.session() as session:
        camp = Campana(
            id="test_camp_1",
            nombre="Test",
            fecha_inicio=date.today(),
            fecha_fin=date.today() + timedelta(days=60),
            estado=EstadoCampana.ACTIVA.value,
            archivo_origen="test.xlsx",
            total_clientes=1,
            total_secciones=1,
        )
        session.add(camp)
        session.commit()

        c = Cliente(
            campana_id="test_camp_1",
            codigo_cliente="C001",
            nombre_completo="Test Client",
            seccion="A",
            nivel_1="Contacto efectivo",
            nivel_2="Promesa de pago",
            nivel_3="Promesa total",
            nivel_4="CAM Promesa total",
            canal_gestion="CAM",
            fecha_promesa_pago="2026-04-15",
            monto_promesa_pago=150.50,
        )
        session.add(c)
        session.commit()
        session.refresh(c)

        assert c.nivel_1 == "Contacto efectivo"
        assert c.nivel_4 == "CAM Promesa total"
        assert c.canal_gestion == "CAM"
        assert c.monto_promesa_pago == 150.50

    cleanup_db(svc, tmp)

test("Cliente nivel fields CRUD", test_db_cliente_nivel_fields)

def test_sync_log_crud():
    from services.database import SyncLog
    svc, tmp = fresh_db()

    with svc.session() as session:
        log = SyncLog(
            tipo="sync_visits",
            registros_afectados=42,
            resultado="ok",
            detalle="test sync",
        )
        session.add(log)
        session.commit()
        session.refresh(log)

        assert log.id is not None
        assert log.tipo == "sync_visits"
        assert log.registros_afectados == 42

    cleanup_db(svc, tmp)

test("SyncLog CRUD", test_sync_log_crud)

# ═══════════════════════════════════════════════════════════════
# 2. CATALOGO NIVELES JSON
# ═══════════════════════════════════════════════════════════════
print("\n=== 2. CATALOGO NIVELES ===")

def test_catalogo_json():
    path = os.path.join(os.path.dirname(__file__), "data", "catalogo_niveles_PE.json")
    assert os.path.exists(path), f"File not found: {path}"
    with open(path, "r", encoding="utf-8") as f:
        cat = json.load(f)
    assert cat["version"] == 1
    assert cat["pais"] == "PE"
    assert "CAM" in cat["canales"]
    assert "TEL" in cat["canales"]
    niveles = cat["niveles"]
    assert len(niveles) > 50, f"Expected 50+ niveles, got {len(niveles)}"

    # Check all entries have required keys
    for i, n in enumerate(niveles):
        for key in ["nivel1", "nivel2", "nivel3", "nivel4", "canal"]:
            assert key in n, f"Entry {i} missing key: {key}"
        assert n["canal"] in ("CAM", "TEL"), f"Entry {i} invalid canal: {n['canal']}"

    # Check unique nivel4 values
    n4_vals = [n["nivel4"] for n in niveles]
    assert len(n4_vals) == len(set(n4_vals)), "Duplicate nivel4 values found"

    # Check both channels have entries
    cam_count = sum(1 for n in niveles if n["canal"] == "CAM")
    tel_count = sum(1 for n in niveles if n["canal"] == "TEL")
    assert cam_count > 20, f"Too few CAM entries: {cam_count}"
    assert tel_count > 20, f"Too few TEL entries: {tel_count}"
    print(f"    Catalog: {len(niveles)} entries ({cam_count} CAM, {tel_count} TEL)")

test("Catalogo JSON valid structure", test_catalogo_json)

def test_catalogo_cascading_logic():
    """Simulate cascading filter logic like the frontend does."""
    path = os.path.join(os.path.dirname(__file__), "data", "catalogo_niveles_PE.json")
    with open(path, "r", encoding="utf-8") as f:
        cat = json.load(f)
    niveles = cat["niveles"]

    # Filter by CAM
    cam = [n for n in niveles if n["canal"] == "CAM"]
    n1_opts = sorted(set(n["nivel1"] for n in cam))
    assert len(n1_opts) >= 2, f"Expected >=2 nivel1 opts for CAM, got {n1_opts}"

    # Pick "Contacto efectivo"
    ce = [n for n in cam if n["nivel1"] == "Contacto efectivo"]
    n2_opts = sorted(set(n["nivel2"] for n in ce))
    assert len(n2_opts) >= 3, f"Expected >=3 nivel2 opts, got {n2_opts}"

    # Pick "Promesa de pago"
    pp = [n for n in ce if n["nivel2"] == "Promesa de pago"]
    n3_opts = sorted(set(n["nivel3"] for n in pp))
    assert len(n3_opts) >= 1

    # Pick first n3
    n3_val = n3_opts[0]
    final = [n for n in pp if n["nivel3"] == n3_val]
    n4_opts = sorted(set(n["nivel4"] for n in final))
    assert len(n4_opts) >= 1
    print(f"    Cascade: CAM > Contacto efectivo > Promesa de pago > {n3_val} -> {n4_opts}")

test("Catalogo cascading filter logic", test_catalogo_cascading_logic)

# ═══════════════════════════════════════════════════════════════
# 3. EXCEL EXPORTER
# ═══════════════════════════════════════════════════════════════
print("\n=== 3. EXCEL EXPORTER ===")

def test_excel_export_basic():
    from services.excel_exporter import export_gestion_excel

    clientes = [
        {
            "codigo_cliente": "C001",
            "nombre_completo": "Juan Pérez",
            "estado_gestion": "visitado_habido",
            "fecha_gestion": "2026-04-05T10:30:00",
            "nivel_1": "Contacto efectivo",
            "nivel_2": "Promesa de pago",
            "nivel_3": "Promesa total",
            "nivel_4": "CAM Promesa total",
            "fecha_promesa_pago": "2026-04-15",
            "monto_promesa_pago": 250.00,
            "nota_gestor": "Cliente se compromete a pagar",
        },
        {
            "codigo_cliente": "C002",
            "nombre_completo": "María García",
            "estado_gestion": "pendiente",
            "fecha_gestion": None,
            "nivel_1": None,
            "nivel_2": None,
            "nivel_3": None,
            "nivel_4": None,
        },
        {
            "codigo_cliente": "C003",
            "nombre_completo": "Carlos López",
            "estado_gestion": "visitado_no_habido",
            "fecha_gestion": "2026-04-06 14:20:00",
            "nivel_1": "No contacto",
            "nivel_2": "No contacto manana / tarde",
            "nivel_3": "No contesta / Mensaje en grabadora",
            "nivel_4": "CAM No gestionada",
        },
    ]

    tmp = tempfile.mktemp(suffix=".xlsx")
    result = export_gestion_excel(
        clientes=clientes,
        output_path=tmp,
        nombre_proveedor="PERECAUDOL",
        solo_gestionados=True,
    )
    assert os.path.exists(result), f"Export file not created: {result}"
    size = os.path.getsize(result)
    assert size > 1000, f"Export file too small: {size} bytes"
    print(f"    Exported: {size} bytes, solo_gestionados=True")

    # Verify content with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    # Header row = 1, data starts at 2
    assert ws.cell(1, 1).value == "Etapa", f"Header A1 != Etapa: {ws.cell(1,1).value}"
    assert ws.cell(1, 2).value == "Nombre proveedor"
    # Only 2 rows (filtered out pendiente C002)
    data_rows = ws.max_row - 1
    assert data_rows == 2, f"Expected 2 data rows (solo_gestionados), got {data_rows}"
    # Check first data row
    assert ws.cell(2, 4).value == "C001", f"Row 2 codigo != C001: {ws.cell(2,4).value}"
    wb.close()
    os.unlink(result)

test("Excel export with solo_gestionados filter", test_excel_export_basic)

def test_excel_export_all():
    from services.excel_exporter import export_gestion_excel

    clientes = [
        {"codigo_cliente": "C001", "estado_gestion": "visitado_habido",
         "fecha_gestion": "2026-04-05T10:30:00", "nivel_1": "Contacto efectivo",
         "nivel_2": "X", "nivel_3": "Y", "nivel_4": "Z"},
        {"codigo_cliente": "C002", "estado_gestion": "pendiente"},
    ]

    tmp = tempfile.mktemp(suffix=".xlsx")
    result = export_gestion_excel(
        clientes=clientes, output_path=tmp,
        nombre_proveedor="TEST", solo_gestionados=False,
    )
    import openpyxl
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    data_rows = ws.max_row - 1
    assert data_rows == 2, f"Expected 2 rows (all), got {data_rows}"
    wb.close()
    os.unlink(result)

test("Excel export all clients (no filter)", test_excel_export_all)

def test_excel_export_empty():
    from services.excel_exporter import export_gestion_excel
    tmp = tempfile.mktemp(suffix=".xlsx")
    result = export_gestion_excel(
        clientes=[], output_path=tmp,
        nombre_proveedor="TEST", solo_gestionados=True,
    )
    assert os.path.exists(result)
    os.unlink(result)

test("Excel export with empty client list", test_excel_export_empty)

# ═══════════════════════════════════════════════════════════════
# 4. CAMPAIGN MANAGER NEW METHODS
# ═══════════════════════════════════════════════════════════════
print("\n=== 4. CAMPAIGN MANAGER ===")

def test_cm_record_sync_and_get_last():
    from services.campaign_manager import CampaignManager
    svc, tmp = fresh_db()

    cm = CampaignManager()
    cm.db = svc

    # Should return None when no syncs
    last = cm.get_last_sync()
    assert last is None, f"Expected None, got {last}"

    # Record one
    cm._record_sync("sync_visits", 10, "ok", "test detail")
    last = cm.get_last_sync()
    assert last is not None
    assert last["tipo"] == "sync_visits"
    assert last["registros_afectados"] == 10
    assert last["resultado"] == "ok"
    print(f"    Last sync: {last}")

    cleanup_db(svc, tmp)

test("CampaignManager _record_sync + get_last_sync", test_cm_record_sync_and_get_last)

def test_cm_cliente_to_dict_nivel_fields():
    from services.database import Cliente, Campana, EstadoCampana
    from services.campaign_manager import CampaignManager
    svc, tmp = fresh_db()

    cm = CampaignManager()
    cm.db = svc

    with svc.session() as session:
        camp = Campana(
            id="tc2", nombre="T", fecha_inicio=date.today(),
            fecha_fin=date.today()+timedelta(days=60),
            estado=EstadoCampana.ACTIVA.value,
            archivo_origen="t.xlsx", total_clientes=1, total_secciones=1,
        )
        session.add(camp)
        session.commit()

        c = Cliente(
            campana_id="tc2", codigo_cliente="X001",
            nombre_completo="Test", seccion="B",
            nivel_1="Contacto efectivo", nivel_2="Promesa de pago",
            nivel_3="Promesa total", nivel_4="CAM Promesa total",
            canal_gestion="CAM", fecha_promesa_pago="2026-04-20",
            monto_promesa_pago=300.0,
        )
        session.add(c)
        session.commit()

        d = cm._cliente_to_dict(c)
        assert d["nivel_1"] == "Contacto efectivo"
        assert d["nivel_4"] == "CAM Promesa total"
        assert d["canal_gestion"] == "CAM"
        assert d["monto_promesa_pago"] == 300.0
        assert d["fecha_promesa_pago"] == "2026-04-20"
        print(f"    Dict keys with niveles: {[k for k in d if 'nivel' in k or 'canal' in k or 'promesa' in k]}")

    cleanup_db(svc, tmp)

test("CampaignManager _cliente_to_dict with nivel fields", test_cm_cliente_to_dict_nivel_fields)

def test_cm_firebase_client_fields():
    from services.campaign_manager import CampaignManager
    # FIREBASE_CLIENT_FIELDS should include new fields
    cm = CampaignManager()
    fields = cm.FIREBASE_CLIENT_FIELDS if hasattr(cm, 'FIREBASE_CLIENT_FIELDS') else []
    # Check as class attribute or module-level
    from services import campaign_manager as cm_mod
    if hasattr(cm_mod, 'FIREBASE_CLIENT_FIELDS'):
        fields = cm_mod.FIREBASE_CLIENT_FIELDS
    
    for f in ["nivel_1", "nivel_2", "nivel_3", "nivel_4", 
              "canal_gestion", "fecha_promesa_pago", "monto_promesa_pago"]:
        assert f in fields, f"Missing field in FIREBASE_CLIENT_FIELDS: {f}"

test("FIREBASE_CLIENT_FIELDS includes nivel fields", test_cm_firebase_client_fields)

def test_cm_restore_campaign_validation():
    """Test restore_campaign_from_firebase with empty data raises ValueError."""
    from services.campaign_manager import CampaignManager
    svc, tmp = fresh_db()

    cm = CampaignManager()
    cm.db = svc

    try:
        cm.restore_campaign_from_firebase({"by_seccion": {}, "metadata": {}})
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "No hay clientes" in str(e)

    cleanup_db(svc, tmp)

test("restore_campaign_from_firebase rejects empty data", test_cm_restore_campaign_validation)

# ═══════════════════════════════════════════════════════════════
# 5. UI PAGES (structural checks)
# ═══════════════════════════════════════════════════════════════
print("\n=== 5. UI PAGES ===")

def test_export_page_structure():
    from ui.pages.export import ExportPage
    assert hasattr(ExportPage, 'render')
    assert hasattr(ExportPage, 'stop')
    assert hasattr(ExportPage, '_on_export')

test("ExportPage has required methods", test_export_page_structure)

def test_sync_page_structure():
    from ui.pages.sync import SyncPage
    assert hasattr(SyncPage, 'render')
    assert hasattr(SyncPage, 'stop')
    assert hasattr(SyncPage, '_on_sync_visits')
    assert hasattr(SyncPage, '_on_upload_catalog')
    assert hasattr(SyncPage, '_on_restore')

test("SyncPage has required methods", test_sync_page_structure)

def test_app_nav_registration():
    with open("ui/app.py", "r", encoding="utf-8") as f:
        content = f.read()
    checks = [
        ('"export"', "NAV_ITEMS export"),
        ('"sync"', "NAV_ITEMS sync"),
        ("ExportPage", "ExportPage import"),
        ("SyncPage", "SyncPage import"),
    ]
    for needle, desc in checks:
        assert needle in content, f"Missing in app.py: {desc}"

test("app.py nav registration", test_app_nav_registration)

# ═══════════════════════════════════════════════════════════════
# 6. FIRESTORE RULES
# ═══════════════════════════════════════════════════════════════
print("\n=== 6. FIRESTORE RULES ===")

def test_firestore_rules():
    rules_path = os.path.join(os.path.dirname(__file__), "..", "firestore.rules")
    with open(rules_path, "r", encoding="utf-8") as f:
        rules = f.read()
    assert "configuracion" in rules, "Missing configuracion rule"
    assert "catalogo_niveles" in rules or "{docId}" in rules

test("Firestore rules include configuracion", test_firestore_rules)

# ═══════════════════════════════════════════════════════════════
# SUMMARY
# ═══════════════════════════════════════════════════════════════
print(f"\n{'='*60}")
print(f"RESULTS: {PASS} passed, {FAIL} failed out of {PASS+FAIL} tests")
print(f"{'='*60}")

if FAIL > 0:
    sys.exit(1)
else:
    print("\nALL TESTS PASSED!")
    sys.exit(0)
